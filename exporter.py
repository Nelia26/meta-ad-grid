from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from models import Config


def export_xlsx(cfg: Config, path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AdGrid"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="A6A6A6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = f"Рекламна сітка (Meta Ads) — {cfg.project or 'Проєкт'}"
    ws.merge_cells("A1:I1")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    headers = [
        "Етап воронки / Ціль",
        "Аудиторія",
        "Таргетинг",
        "Офер / УТП",
        "Креатив",
        "Плейсменти",
        "Оптимізація / Подія",
        "KPI",
        "Бюджет",
    ]
    ws.append([""] * 9)  # row 2 empty
    ws.append([""] * 9)  # row 3 empty
    ws.append(headers)  # row 4

    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for i, row in enumerate(cfg.rows, start=5):
        values = [
            row.stage,
            row.audience,
            row.targeting,
            row.offer,
            row.creative,
            row.placements,
            row.optimization_event,
            row.kpi,
            row.budget,
        ]
        ws.append(values)
        for col in range(1, 10):
            c = ws.cell(row=i, column=col)
            c.alignment = wrap
            c.border = border

    widths = [22, 24, 34, 22, 26, 18, 22, 18, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    ws.row_dimensions[4].height = 36

    wb.save(path)
